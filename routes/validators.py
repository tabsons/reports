from flask import Flask,  Blueprint, request, render_template, send_file, Response
import pandas as pd
import os
import xlrd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import re
import zipfile
import os
import numpy as np
import torch
from openpyxl.styles import Color, PatternFill, Font, Border, NamedStyle
from scipy.spatial.distance import cosine
from sentence_transformers import SentenceTransformer
from datetime import time
from openpyxl.styles import NamedStyle
from datetime import timedelta


val_route = Blueprint('val_route', __name__)

def j_s(s1, s2):
    set1 = set(s1)
    set2 = set(s2)
    
    intersection = set1.intersection(set2)
    union = set1.union(set2)
    
    if not union:  # Avoid division by zero
        return 0.0
    
    similarity = len(intersection) / len(union)
    return similarity

def r_s_t(text, threshold, target_string):
    text_without_space = text.replace(" ", "").upper()
    target_without_space = target_string.replace(" ", "").upper()  #ADD THIS LINE
    similarity = j_s(text_without_space, target_string)
    
    len_difference = len(text_without_space) - len(target_without_space) #ADD THIS LINE
    if similarity > threshold and len_difference <= 3: #ADD THIS LINE
        return "BAD RECORDING"
    elif re.search(r"\bOTHER LANGUAGE\b",text.upper()):
        return "OTHER LANGUAGE"
    elif re.search(r"\bANOTHER LANGUAGE\b",text.upper()):
        return "OTHER LANGUAGE"
    return text  # Return the original text if no conditions are met

def remove_special_characters(text):

    pattern = r'[^a-zA-Z0-9\s./]+'


    cleaned_text = re.sub(pattern, ' ', text)

    end_dot_pattern = r'\.(?!\S)'

    cleaned_text = re.sub(end_dot_pattern, ' ', cleaned_text)

    # Replace multiple spaces with a single space
    cleaned_text = re.sub(r'\s+', ' ', cleaned_text).strip()

    return cleaned_text


@val_route.route('/validations/test/', methods=['GET', 'POST'])
def test_validation_report():
    if request.method == 'POST':
        file = request.files['file']
        df = pd.read_excel(file, sheet_name="expressreport")

        # df.iloc[:, 13] = pd.to_datetime(df.iloc[:, 13], errors='coerce')
        # df.iloc[:, 17] = pd.to_datetime(df.iloc[:, 17], errors='coerce')


        # for x in range(len(df) - 1):
        #     if df.iloc[x + 1, 13] != df.iloc[x, 17]:
        #         df.iloc[x + 1, 13] = df.iloc[x, 17]

        # df.iloc[:, 13] = pd.to_datetime(df.iloc[:, 13]).dt.strftime('%H:%M:%S')
        # df.iloc[:, 17] = pd.to_datetime(df.iloc[:, 17]).dt.strftime('%H:%M:%S')
        df['Sub-Story'] = df['Sub-Story'].astype(str)
        df['Sub-Story'] = df['Sub-Story'].apply(remove_special_characters)
        df['Story'] = df['Story'].astype(str)
        df['Story'] = df['Story'].apply(remove_special_characters)

        if (df.notna().sum() > 1).any():
            pass
        else:
            return "Please, Correct the file. No fields.", 400
        target_string = "BADRECORDING"
        threshold = .72
        # df["Story"] = df["Story"].apply(lambda x: r_s_t(str(x) , threshold , target_string))
        # df["Sub-Story"] = df["Sub-Story"].apply(lambda x: r_s_t(str(x) , threshold , target_string))

        headers = pd.read_excel('files/original_headers.xlsx')
        if len(df.columns) != len(headers.columns):
            print(len(df.columns),len(headers.columns) )
            return Response('You have added or removed some headers in your file. Please correct and try again.')
        for col in df.columns:
            if col not in headers.columns:
                return Response('Please check the headers')
        if 'Sr No' in df.columns:
            df.loc[((df['Story']=='ANCHOR INTRODUCTION') & (df['Telecast Format'].str.lower()!='shows')),'Telecast Format'] = 'NO NEWS CONTENT'
            GA1 = df[(df['Story Format']=='DEBATE OR DISCUSSION')& (df['Guest'].isna() | df['Anchor'].isna())]
            PA = df[(df['Story Format']=='INTERVIEW')& (df['Personality'].isna() | df['Anchor'].isna())]
            A = df[(df['Telecast Format'].str.lower() =='shows') & (df['Anchor'].isna())]
            GA2 = df[(df['Story Format']=='TELEPHONIC CONVERSATION')& (df['Guest'].isna() | df['Anchor'].isna())]
            GA = pd.concat([GA1,GA2])
            grouped = df.groupby('Story')['Story Genre'].agg(list)
            duplicates= grouped[grouped.apply(lambda x: len(set(x))>1)]
            group = df.groupby('Story')['Story Genre'].agg(set)
            dupli= group[group.apply(lambda x: len(set(x))>1)]
            new_df = pd.DataFrame(dupli).reset_index()
            print(len(GA1) + len(GA2))
            print(len(GA))
            lst=[]
            count_dict={}

            for i,x in enumerate(duplicates):
                for y in set(x):
                    lst.append(f"{y} : {x.count(y)}")
                    count_dict[y] = x.count(y)
                new_df.loc[i,'Story Genre Count'] = ' '.join(lst)
                new_df.loc[i, 'Story Genre']= max(count_dict, key=count_dict.get)
                lst.clear()
                count_dict.clear()
            merged_df = pd.merge(df, new_df, on='Story' , how='left')
            # df['Story Genre Suggestion'] = merged_df['Story Genre_y']
            df.insert(loc=11, column='Story Genre Suggestion', value=merged_df['Story Genre_y'])

            new_df.drop('Story Genre',axis=1,  inplace=True)
            geography = df.groupby('Sub-Story')['Geography'].agg(list)
            geog_duplicates= geography[geography.apply(lambda x: len(set(x))>1)]
            geog = df.groupby('Sub-Story')['Geography'].agg(set)
            geog_dup= geog[geog.apply(lambda x: len(set(x))>1)]
            geog_df = pd.DataFrame(geog_dup).reset_index()
            lst2=[]
            count_dict2={}
            for i,x in enumerate(geog_duplicates):

                for y in set(x):

                    # print(y , x.count(y))
                    lst2.append(f"{y} : {x.count(y)}")
                    count_dict2[y] = x.count(y)
                geog_df.loc[i,'Geography Count'] = ' '.join(lst2)
                geog_df.loc[i, 'Geography']= max(count_dict2, key=count_dict2.get)
                lst2.clear()
                count_dict2.clear()

            merged_df = pd.merge(df, geog_df, on='Sub-Story' , how='left')
            # df['Geography Suggestion'] = merged_df['Geography_y']
            df.insert(loc=24, column='Geography Suggestion', value=merged_df['Geography_y'])
            
            story = df.groupby('Sub-Story')['Story'].agg(list)
            story_duplicates= story[story.apply(lambda x: len(set(x))>1)]
            story_g = df.groupby('Sub-Story')['Story'].agg(set)
            story_dupli= story_g[story_g.apply(lambda x: len(set(x))>1)]
            story_new = pd.DataFrame(story_dupli).reset_index()
            lst3=[]
            count_dict3={}
            for i,x in enumerate(story_duplicates):

                for y in set(x):

                    # print(y , x.count(y))
                    lst3.append(f"{y} : {x.count(y)}")
                    count_dict3[y] = x.count(y)
                story_new.loc[i,'Story Count'] = ' '.join(lst3)
                story_new.loc[i, 'Suggested Story']= max(count_dict3, key=count_dict3.get)
                lst3.clear()
                count_dict3.clear()
            merged_df = pd.merge(df, story_new, on='Sub-Story' , how='left')
            
            # df['Story Suggestion'] = merged_df['Suggested Story']
            if len(story_new) == 0:
                df.insert(loc=9, column='Story Suggestion', value=merged_df['Story_y'])
            else:
                df.insert(loc=9, column='Story Suggestion', value=merged_df['Suggested Story'])
            # excel_writer = pd.ExcelWriter('validations.xlsx', engine='openpyxl')
            # new_df.to_excel(excel_writer, sheet_name='expressreport', index=False)
            # new_df.to_excel(excel_writer, sheet_name='Multiple Story Genre', index=False)
            # geog_df.to_excel(excel_writer, sheet_name='Multiple Geography', index=False)
            # GA.to_excel(excel_writer, sheet_name='Guest Anchor Nulls', index=False)
            # A.to_excel(excel_writer, sheet_name='Anchor Nulls', index=False)
            # PA.to_excel(excel_writer, sheet_name='Anchor Personality Nulls', index=False)
            # excel_writer.close()
            story_p = df.groupby('Sub-Story')['Personality'].agg(list)
            personality_duplicates= story_p[story_p.apply(lambda x: len(set(x))>1)]
            story_per = df.groupby('Sub-Story')['Personality'].agg(set)

            personality_dupli= story_per[story_per.apply(lambda x: len(set(x))>1)]
            substory_new = pd.DataFrame(personality_dupli).reset_index()
            lst4=[]
            count_dict4={}
            if len(personality_duplicates) > 0:
                for i,x in enumerate(personality_duplicates):

                    for y in set(x):

                        # print(y , x.count(y))
                        lst4.append(f"{y} : {x.count(y)}")
                        count_dict4[y] = x.count(y)
                    substory_new.loc[i,'Personality Count'] = ' '.join(lst4)

                    count_dict4 = {k: v for k, v in count_dict4.items() if k == k}

                    substory_new.loc[i, 'Suggested Personality']= max(count_dict4, key=count_dict4.get)

                    # print(sorted(count_dict3, key=count_dict3.get)[-2])
                    lst4.clear()
                    count_dict4.clear()

            merged_df = pd.merge(df, substory_new, on='Sub-Story' , how='left')
            if len(personality_duplicates) > 0:
                # df['Personality Suggestion'] = merged_df['Suggested Personality']
                df.insert(loc=34, column='Personality Suggestion', value=merged_df['Suggested Personality'])

            df.loc[df["Story"]=="BAD RECORDING",[
                "Story Genre",
                "Geography",
                "Logistics",
                "Split",
                "Story Format",
                "Event/Conclave",
                "LIVE status",
                "Exclusive",
                "Telecast Format",
                "Personality",
                "Guest",
                "Anchor",
                "Reporter"
            ]] = [
                "NO NEWS CONTENT",
                "NA",
                "NA",
                "NA",
                "NO NEWS CONTENT",
                "NA",
                "NA",
                "NA",
                "NO NEWS CONTENT",
                "",
                "",
                "",
                ""

            ]
            

            wb = load_workbook('files\\header_copy.xlsx')
            sheet_name = 'expressreport' 
            make_file(wb, 'expressreport', df, False)
            make_file(wb, 'Multiple Story Genre', new_df, True)
            make_file(wb, 'Multiple Geography', geog_df, True)
            make_file(wb, 'Guest Anchor Nulls', GA, True)
            make_file(wb, 'Anchor Nulls', A, True)
            make_file(wb, 'Anchor Personality Nulls', PA, True)
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return send_file(output,download_name='expressreport.xlsx', as_attachment=True)
        else:
            message = 'Please add Sr No as the first column in file and then try again.'
            return render_template('33_validations.html', message = message)
    return render_template('33_validations.html')

@val_route.route('/validations/', methods=['GET', 'POST'])
def validations():
    if request.method == 'POST':
        file = request.files['file']
        df = pd.read_excel(file, sheet_name="expressreport")

        # df.iloc[:, 13] = pd.to_datetime(df.iloc[:, 13], errors='coerce')
        # df.iloc[:, 17] = pd.to_datetime(df.iloc[:, 17], errors='coerce')


        # for x in range(len(df) - 1):
        #     if df.iloc[x + 1, 13] != df.iloc[x, 17]:
        #         df.iloc[x + 1, 13] = df.iloc[x, 17]

        # df.iloc[:, 13] = pd.to_datetime(df.iloc[:, 13]).dt.strftime('%H:%M:%S')
        # df.iloc[:, 17] = pd.to_datetime(df.iloc[:, 17]).dt.strftime('%H:%M:%S')
        df['Sub-Story'] = df['Sub-Story'].astype(str)
        df['Sub-Story'] = df['Sub-Story'].apply(remove_special_characters)
        df['Story'] = df['Story'].astype(str)
        df['Story'] = df['Story'].apply(remove_special_characters)

        if (df.notna().sum() > 1).any():
            pass
        else:
            return "Please, Correct the file. No fields.", 400
        target_string = "BADRECORDING"
        threshold = .72
        # df["Story"] = df["Story"].apply(lambda x: r_s_t(str(x) , threshold , target_string))
        # df["Sub-Story"] = df["Sub-Story"].apply(lambda x: r_s_t(str(x) , threshold , target_string))

        headers = pd.read_excel('files/original_headers.xlsx')
        if len(df.columns) != len(headers.columns):
            print(len(df.columns),len(headers.columns) )
            return Response('You have added or removed some headers in your file. Please correct and try again.')
        for col in df.columns:
            if col not in headers.columns:
                return Response('Please check the headers')
        if 'Sr No' in df.columns:
            df.loc[((df['Story']=='ANCHOR INTRODUCTION') & (df['Telecast Format'].str.lower()!='shows')),'Telecast Format'] = 'NO NEWS CONTENT'
            GA1 = df[(df['Story Format']=='DEBATE OR DISCUSSION')& (df['Guest'].isna() | df['Anchor'].isna())]
            PA = df[(df['Story Format']=='INTERVIEW')& (df['Personality'].isna() | df['Anchor'].isna())]
            A = df[(df['Telecast Format'].str.lower() =='shows') & (df['Anchor'].isna())]
            GA2 = df[(df['Story Format']=='TELEPHONIC CONVERSATION')& (df['Guest'].isna() | df['Anchor'].isna())]
            GA = pd.concat([GA1,GA2])
            grouped = df.groupby('Story')['Story Genre'].agg(list)
            duplicates= grouped[grouped.apply(lambda x: len(set(x))>1)]
            group = df.groupby('Story')['Story Genre'].agg(set)
            dupli= group[group.apply(lambda x: len(set(x))>1)]
            new_df = pd.DataFrame(dupli).reset_index()
            print(len(GA1) + len(GA2))
            print(len(GA))
            lst=[]
            count_dict={}

            for i,x in enumerate(duplicates):
                for y in set(x):
                    lst.append(f"{y} : {x.count(y)}")
                    count_dict[y] = x.count(y)
                new_df.loc[i,'Story Genre Count'] = ' '.join(lst)
                new_df.loc[i, 'Story Genre']= max(count_dict, key=count_dict.get)
                lst.clear()
                count_dict.clear()
            merged_df = pd.merge(df, new_df, on='Story' , how='left')
            # df['Story Genre Suggestion'] = merged_df['Story Genre_y']
            df.insert(loc=11, column='Story Genre Suggestion', value=merged_df['Story Genre_y'])

            new_df.drop('Story Genre',axis=1,  inplace=True)
            geography = df.groupby('Sub-Story')['Geography'].agg(list)
            geog_duplicates= geography[geography.apply(lambda x: len(set(x))>1)]
            geog = df.groupby('Sub-Story')['Geography'].agg(set)
            geog_dup= geog[geog.apply(lambda x: len(set(x))>1)]
            geog_df = pd.DataFrame(geog_dup).reset_index()
            lst2=[]
            count_dict2={}
            for i,x in enumerate(geog_duplicates):

                for y in set(x):

                    # print(y , x.count(y))
                    lst2.append(f"{y} : {x.count(y)}")
                    count_dict2[y] = x.count(y)
                geog_df.loc[i,'Geography Count'] = ' '.join(lst2)
                geog_df.loc[i, 'Geography']= max(count_dict2, key=count_dict2.get)
                lst2.clear()
                count_dict2.clear()

            merged_df = pd.merge(df, geog_df, on='Sub-Story' , how='left')
            # df['Geography Suggestion'] = merged_df['Geography_y']
            df.insert(loc=23, column='Geography Suggestion', value=merged_df['Geography_y'])
            
            story = df.groupby('Sub-Story')['Story'].agg(list)
            story_duplicates= story[story.apply(lambda x: len(set(x))>1)]
            story_g = df.groupby('Sub-Story')['Story'].agg(set)
            story_dupli= story_g[story_g.apply(lambda x: len(set(x))>1)]
            story_new = pd.DataFrame(story_dupli).reset_index()
            lst3=[]
            count_dict3={}
            for i,x in enumerate(story_duplicates):

                for y in set(x):

                    # print(y , x.count(y))
                    lst3.append(f"{y} : {x.count(y)}")
                    count_dict3[y] = x.count(y)
                story_new.loc[i,'Story Count'] = ' '.join(lst3)
                story_new.loc[i, 'Suggested Story']= max(count_dict3, key=count_dict3.get)
                lst3.clear()
                count_dict3.clear()
            merged_df = pd.merge(df, story_new, on='Sub-Story' , how='left')
            
            # df['Story Suggestion'] = merged_df['Suggested Story']
            if len(story_new) == 0:
                df.insert(loc=9, column='Story Suggestion', value=merged_df['Story_y'])
            else:
                df.insert(loc=9, column='Story Suggestion', value=merged_df['Suggested Story'])
            # excel_writer = pd.ExcelWriter('validations.xlsx', engine='openpyxl')
            # new_df.to_excel(excel_writer, sheet_name='expressreport', index=False)
            # new_df.to_excel(excel_writer, sheet_name='Multiple Story Genre', index=False)
            # geog_df.to_excel(excel_writer, sheet_name='Multiple Geography', index=False)
            # GA.to_excel(excel_writer, sheet_name='Guest Anchor Nulls', index=False)
            # A.to_excel(excel_writer, sheet_name='Anchor Nulls', index=False)
            # PA.to_excel(excel_writer, sheet_name='Anchor Personality Nulls', index=False)
            # excel_writer.close()
            story_p = df.groupby('Sub-Story')['Personality'].agg(list)
            personality_duplicates= story_p[story_p.apply(lambda x: len(set(x))>1)]
            story_per = df.groupby('Sub-Story')['Personality'].agg(set)

            personality_dupli= story_per[story_per.apply(lambda x: len(set(x))>1)]
            substory_new = pd.DataFrame(personality_dupli).reset_index()
            lst4=[]
            count_dict4={}
            if len(personality_duplicates) > 0:
                for i,x in enumerate(personality_duplicates):

                    for y in set(x):

                        # print(y , x.count(y))
                        lst4.append(f"{y} : {x.count(y)}")
                        count_dict4[y] = x.count(y)
                    substory_new.loc[i,'Personality Count'] = ' '.join(lst4)

                    count_dict4 = {k: v for k, v in count_dict4.items() if k == k}

                    substory_new.loc[i, 'Suggested Personality']= max(count_dict4, key=count_dict4.get)

                    # print(sorted(count_dict3, key=count_dict3.get)[-2])
                    lst4.clear()
                    count_dict4.clear()

            merged_df = pd.merge(df, substory_new, on='Sub-Story' , how='left')
            if len(personality_duplicates) > 0:
                # df['Personality Suggestion'] = merged_df['Suggested Personality']
                df.insert(loc=33, column='Personality Suggestion', value=merged_df['Suggested Personality'])

            df.loc[df["Story"]=="BAD RECORDING",[
                "Story Genre",
                "Geography",
                "Logistics",
                "Split",
                "Story Format",
                "Event/Conclave",
                "LIVE status",
                "Exclusive",
                "Telecast Format",
                "Personality",
                "Guest",
                "Anchor",
                "Reporter"
            ]] = [
                "NO NEWS CONTENT",
                "NA",
                "NA",
                "NA",
                "NO NEWS CONTENT",
                "NA",
                "NA",
                "NA",
                "NO NEWS CONTENT",
                "",
                "",
                "",
                ""

            ]
            

            wb = load_workbook('files\\header.xlsx')
            sheet_name = 'expressreport' 
            make_file(wb, 'expressreport', df, False)
            make_file(wb, 'Multiple Story Genre', new_df, True)
            make_file(wb, 'Multiple Geography', geog_df, True)
            make_file(wb, 'Guest Anchor Nulls', GA, True)
            make_file(wb, 'Anchor Nulls', A, True)
            make_file(wb, 'Anchor Personality Nulls', PA, True)
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return send_file(output,download_name='expressreport.xlsx', as_attachment=True)
        else:
            message = 'Please add Sr No as the first column in file and then try again.'
            return render_template('33_validations.html', message = message)
    return render_template('33_validations.html')    


@val_route.route('/test-validations', methods=['GET', 'POST'])
def test_validations():
    if request.method == 'POST':
        file = request.files['file']
        df = pd.read_excel(file, sheet_name="expressreport")

        df = df.sort_values(by=['Channel Name', 'Program Date', 'Clip Start Time'], ascending=[True, True, True])

        # Check if there are any blanks in column 13 (Clip Start Time)
        if df.iloc[:, 13].isna().any() or (df.iloc[:, 13] == "").any():
            return "Blank exists in Clip Start Time (Column 13)", 400

        # Check if there are any blanks in column 17 (Clip End Time)
        if df.iloc[:, 17].isna().any() or (df.iloc[:, 17] == "").any():
            return "Blank exists in Clip End Time (Column 17)", 400

        df.iloc[:, 13] = pd.to_datetime(df.iloc[:, 13], errors='coerce')
        df.iloc[:, 17] = pd.to_datetime(df.iloc[:, 17], errors='coerce')

        for x in range(len(df) - 1):
            # Drift
            time_diff = (df.iloc[x + 1, 13] - df.iloc[x, 17]).total_seconds()
            if time_diff <= 2.0 and 0 < time_diff:
                df.iloc[x + 1, 13] = df.iloc[x, 17]
            # Overlap
            time_diff1 = (df.iloc[x, 17] - df.iloc[x + 1, 13]).total_seconds()
            if time_diff1 <= 2.0 and 0 < time_diff1:
                df.iloc[x + 1, 13] = df.iloc[x, 17]

        lst = []
        lst1 = []
        for x in range(len(df)):
            print(df.iloc[x, 17], df.iloc[x, 13])
            lst.append(df.iloc[x, 17] - df.iloc[x, 13])
            lst1.append(int((df.iloc[x, 17] - df.iloc[x, 13]).total_seconds()))

        df['new_duration_hours'] = lst
        df['new_duration_seconds'] = lst1
        df['new_time'] = df.iloc[:, 34].dt.total_seconds().apply(
            lambda x: f'{int(x // 3600):02}:{int((x % 3600) // 60):02}:{int(x % 60):02}')
        df['Duration (HH:MM:SS)'] = df['new_time']
        df['Duration (seconds)'] = df['new_duration_seconds']
        df.iloc[:, 13] = pd.to_datetime(df.iloc[:, 13]).dt.strftime('%H:%M:%S')
        df.iloc[:, 17] = pd.to_datetime(df.iloc[:, 17]).dt.strftime('%H:%M:%S')
        df.drop(columns=['new_duration_hours', 'new_duration_seconds', 'new_time'], inplace=True)
        df['key'] = df['Program Date'].astype(str) + '_' + df['Channel Code'].astype(str) + '_' + df[
            'Clip Start Time'].astype(str)
        df['key'] = df['key'].where(df['key'].duplicated(keep=False), 0)
        # drift correctinon end

        df['Sub-Story'] = df['Sub-Story'].astype(str)
        df['Sub-Story'] = df['Sub-Story'].apply(remove_special_characters)
        df['Story'] = df['Story'].astype(str)
        df['Story'] = df['Story'].apply(remove_special_characters)

        # correcting other language and bad recording
        if (df.notna().sum() > 1).any():
            pass
        else:
            return "Please, Correct the file. No fields.", 400
        target_string = "badrecording"
        threshold = .8
        # other language
        # df["Story"] = df["Story"].apply(lambda x: r_s_t(str(x) , threshold , target_string))


        headers = pd.read_excel('files/original_headers.xlsx')
        # if len(df.columns) != len(headers.columns):
        #     print(len(df.columns),len(headers.columns) )
        #     return Response('You have added or removed some headers in your file. Please correct and try again.')
        # for col in df.columns:
        #     if col not in headers.columns:
        #         return Response('Please check the headers')
        if 'Sr No' in df.columns:
            df.loc[((df['Story'] == 'ANCHOR INTRODUCTION') & (
                        df['Telecast Format'].str.lower() != 'shows')), 'Telecast Format'] = 'NO NEWS CONTENT'
            GA1 = df[(df['Story Format'] == 'DEBATE OR DISCUSSION') & (df['Guest'].isna() | df['Anchor'].isna())]
            PA = df[(df['Story Format'] == 'INTERVIEW') & (df['Personality'].isna() | df['Anchor'].isna())]
            A = df[(df['Telecast Format'].str.lower() == 'shows') & (df['Anchor'].isna())]
            GA2 = df[(df['Story Format'] == 'TELEPHONIC CONVERSATION') & (df['Guest'].isna() | df['Anchor'].isna())]
            GA = pd.concat([GA1, GA2])
            grouped = df.groupby('Story')['Story Genre'].agg(list)
            duplicates = grouped[grouped.apply(lambda x: len(set(x)) > 1)]
            group = df.groupby('Story')['Story Genre'].agg(set)
            dupli = group[group.apply(lambda x: len(set(x)) > 1)]
            new_df = pd.DataFrame(dupli).reset_index()
            print(len(GA1) + len(GA2))
            print(len(GA))
            lst = []
            count_dict = {}

            for i, x in enumerate(duplicates):
                for y in set(x):
                    lst.append(f"{y} : {x.count(y)}")
                    count_dict[y] = x.count(y)
                new_df.loc[i, 'Story Genre Count'] = ' '.join(lst)
                new_df.loc[i, 'Story Genre'] = max(count_dict, key=count_dict.get)
                lst.clear()
                count_dict.clear()
            merged_df = pd.merge(df, new_df, on='Story', how='left')
            df.insert(loc=11, column='Story Genre Suggestion', value=merged_df['Story Genre_y'])

            new_df.drop('Story Genre', axis=1, inplace=True)
            geography = df.groupby('Sub-Story')['Geography'].agg(list)
            geog_duplicates = geography[geography.apply(lambda x: len(set(x)) > 1)]
            geog = df.groupby('Sub-Story')['Geography'].agg(set)
            geog_dup = geog[geog.apply(lambda x: len(set(x)) > 1)]
            geog_df = pd.DataFrame(geog_dup).reset_index()
            lst2 = []
            count_dict2 = {}
            for i, x in enumerate(geog_duplicates):

                for y in set(x):
                    # print(y , x.count(y))
                    lst2.append(f"{y} : {x.count(y)}")
                    count_dict2[y] = x.count(y)
                geog_df.loc[i, 'Geography Count'] = ' '.join(lst2)
                geog_df.loc[i, 'Geography'] = max(count_dict2, key=count_dict2.get)
                lst2.clear()
                count_dict2.clear()

            merged_df = pd.merge(df, geog_df, on='Sub-Story', how='left')
            df.insert(loc=23, column='Geography Suggestion', value=merged_df['Geography_y'])

            story = df.groupby('Sub-Story')['Story'].agg(list)
            story_duplicates = story[story.apply(lambda x: len(set(x)) > 1)]
            story_g = df.groupby('Sub-Story')['Story'].agg(set)
            story_dupli = story_g[story_g.apply(lambda x: len(set(x)) > 1)]
            story_new = pd.DataFrame(story_dupli).reset_index()
            lst3 = []
            count_dict3 = {}
            for i, x in enumerate(story_duplicates):

                for y in set(x):
                    # print(y , x.count(y))
                    lst3.append(f"{y} : {x.count(y)}")
                    count_dict3[y] = x.count(y)
                story_new.loc[i, 'Story Count'] = ' '.join(lst3)
                story_new.loc[i, 'Suggested Story'] = max(count_dict3, key=count_dict3.get)
                lst3.clear()
                count_dict3.clear()
            merged_df = pd.merge(df, story_new, on='Sub-Story', how='left')

            if len(story_new) == 0:
                df.insert(loc=9, column='Story Suggestion', value=merged_df['Story_y'])
            else:
                df.insert(loc=9, column='Story Suggestion', value=merged_df['Suggested Story'])

            story_p = df.groupby('Sub-Story')['Personality'].agg(list)
            personality_duplicates = story_p[story_p.apply(lambda x: len(set(x)) > 1)]
            story_per = df.groupby('Sub-Story')['Personality'].agg(set)

            personality_dupli = story_per[story_per.apply(lambda x: len(set(x)) > 1)]
            substory_new = pd.DataFrame(personality_dupli).reset_index()
            lst4 = []
            count_dict4 = {}
            if len(personality_duplicates) > 0:
                for i, x in enumerate(personality_duplicates):

                    for y in set(x):
                        # print(y , x.count(y))
                        lst4.append(f"{y} : {x.count(y)}")
                        count_dict4[y] = x.count(y)
                    substory_new.loc[i, 'Personality Count'] = ' '.join(lst4)

                    count_dict4 = {k: v for k, v in count_dict4.items() if k == k}

                    substory_new.loc[i, 'Suggested Personality'] = max(count_dict4, key=count_dict4.get)

                    lst4.clear()
                    count_dict4.clear()

            merged_df = pd.merge(df, substory_new, on='Sub-Story', how='left')
            if len(personality_duplicates) > 0:
                df.insert(loc=33, column='Personality Suggestion', value=merged_df['Suggested Personality'])
            df['Duration (HH:MM:SS)'] = df['Duration (HH:MM:SS)'].apply(
                lambda x: x.strftime('%H:%M:%S') if isinstance(x, time) else x
            )
            df['Duration (HH:MM:SS)'] = pd.to_timedelta(df['Duration (HH:MM:SS)'], errors='coerce')

            df.loc[df["Story"] == "BAD RECORDING", [
                "Story Genre",
                "Geography",
                "Logistics",
                "Split",
                "Story Format",
                "Event/Conclave",
                "LIVE status",
                "Exclusive",
                "Telecast Format",
                "Personality",
                "Guest",
                "Anchor",
                "Reporter"
            ]] = [
                "NO NEWS CONTENT",
                "NA",
                "NA",
                "NA",
                "NO NEWS CONTENT",
                "NA",
                "NA",
                "NA",
                "NO NEWS CONTENT",
                "",
                "",
                "",
                ""

            ]

            wb = load_workbook('files\\header.xlsx')
            sheet_name = 'expressreport'
            ws = wb[sheet_name]

            # Define red fill for the cells
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            # Iterate over the rows in df to apply the red fill conditionally
            for row_idx, row in enumerate(df.itertuples(), start=2):  # Assuming row 1 is the header
                if row.key != 0:  # Check the condition for applying red fill (ensure 'key' is valid)
                    for col_idx, value in enumerate(row[1:], start=1):  # Iterate over all columns in the row
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = value  # Set the value of the cell
                        cell.fill = red_fill  # Apply red fill
                elif row.key == 0:
                    for col_idx, value in enumerate(row[1:], start=1):  # Iterate over all columns in the row
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = value  # Set the value of the cell
            ws.delete_cols(39)
            make_file(wb, 'Multiple Story Genre', new_df, True)
            make_file(wb, 'Multiple Geography', geog_df, True)
            make_file(wb, 'Guest Anchor Nulls', GA, True)
            make_file(wb, 'Anchor Nulls', A, True)
            make_file(wb, 'Anchor Personality Nulls', PA, True)
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return send_file(output, download_name='expressreport.xlsx', as_attachment=True)
        else:
            message = 'Please add Sr No as the first column in file and then try again.'
            return render_template('33_validations.html', message=message)
    return render_template('33_validations.html')


def make_file(wb, sheet_name, df, is_header):    
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        next_row = ws.max_row + 1
    else:
        ws = wb[sheet_name]
        next_row = ws.max_row + 1
    # Append the new data to the selected sheet
    if is_header:
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=is_header), 0):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=next_row+r_idx-1, column=c_idx, value=value)
    else:
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=is_header), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=next_row+r_idx-1, column=c_idx, value=value)

@val_route.route('/qc_log', methods=['GET', 'POST'])
def qc_log_correction():
    if request.method == 'POST':
        file = request.files['file']
        df = pd.read_excel(file, sheet_name='Qc News Log')
        df['Actual Error'] = df['Actual Error'].str.rstrip(',')
        df['Actual Error'] = df['Actual Error'].str.rstrip(' ')
        df_filtered = df[df.apply(lambda row: str_to_set(row['Actual Error']) != str_to_set(row['Correction Version']), axis=1)]
        df_filtered['Clip Start Time'] = df_filtered['Clip Start Time'].dt.time
        df_filtered['Clip End Time'] = df_filtered['Clip End Time'].dt.time
        excel = 'qc_log_corrected.xlsx'
        df_filtered.to_excel(excel, index=False)
        return send_file(excel, as_attachment=True)
    return render_template('qc_log_correction.html')

def str_to_set(s):
    if pd.isna(s):
        return set()
    return set(map(str.strip, s.lower().split(',')))


def get_soundex(word):
    token = word.upper()
    soundex = ""

    soundex += token[0]


    dictionary = {"BFPV": "1", "CGJKQSXZ":"2", "DT":"3", "L":"4", "MN":"5", "R":"6", "AEIOUHWY":"."}
    for char in token[1:]:
        for key in dictionary.keys():
            if char in key:
                code = dictionary[key]
                if code != soundex[-1]:
                    soundex += code

    soundex = soundex.replace(".", "")

    soundex = soundex[:6].ljust(6, "0")

    return soundex

@val_route.route('/validate_logger_data/', methods=['GET', 'POST'])
def logger_data_validation():
    if request.method == 'POST':
        file = request.files['file']
        df = pd.read_excel(file, sheet_name='Sheet1')
        df[(df['TelecastFormat'].str.lower()=='shows') & df['Anchor'].isna()]
        df[(df['StoryFormat']=='DEBATE OR DISCUSSION') & df['Guest'].isna()]
        grouped = df.groupby('Description')['StoryGenre'].agg(list)
        duplicates= grouped[grouped.apply(lambda x: len(set(x))>1)]
        group = df.groupby('Description')['StoryGenre'].agg(set)
        dupli= group[group.apply(lambda x: len(set(x))>1)]
        new_df = pd.DataFrame(dupli).reset_index()
        lst=[]
        count_dict={}
        for i,x in enumerate(duplicates):

            for y in set(x):

                # print(y , x.count(y))
                lst.append(f"{y} : {x.count(y)}")
                count_dict[y] = x.count(y)
            # print(max(count_dict, key=count_dict.get))
            new_df.loc[i,'Story Genre Count'] = ' '.join(lst)
            new_df.loc[i, 'Story Genre Sugession']= max(count_dict, key=count_dict.get)
            lst.clear()
            count_dict.clear()

        merged_df = pd.merge(df, new_df, on='Description' , how='left')
        # df['Story Genre'] = merged_df['Story Genre_y'].fillna(df['Story Genre'])
        df['Story Genre Suggestion'] = merged_df['Story Genre Sugession'] 
        df[(df['ProgramType'].str.lower() != 'filler') &  (df['Description'].isna())]
        df[(df['ProgramType'].str.lower() != 'filler') &  (df['SubStory'].isna())]
  
        excel = 'logger_data_corrected.xlsx'
        df.to_excel(excel, index=False)
        return send_file(excel, as_attachment=True)


    return render_template('log_correction.html')

def create_zip(zip_name, files):
    """
    Creates a zip file containing the specified files.

    Args:
        zip_name (str): The name of the output zip file.
        files (list): List of file paths to include in the zip.

    Returns:
        str: Path to the created zip file.
    """
    with zipfile.ZipFile(zip_name, 'w') as zipf:
        for file in files:
            zipf.write(file, os.path.basename(file))  # Add file with its basename to avoid including full path
    return zip_name


@val_route.route('/excel_to_csv/', methods=['GET', 'POST'])
def excel_to_csv():
    if request.method == 'POST':
        file = request.files['file']
        df = pd.read_excel(file, sheet_name='expressreport')
        column_mapping = {
            "Sr No": "Sr No",
            "Channel Name": "channel",
            "Channel Code": "BARC_CHNCD",
            "Program Date": "pgm_date",
            "Week": "week",
            "Week Day": "week_day",
            "Program Start Time": "Pgm_Start_Time",
            "Program End Time": "Pgm_End_Time",
            "Clip Start Time": "clip_start_time",
            "Clip End Time": "clip_end_time",
            "Story": "Story",
            "Sub-Story": "Sub_Story",
            "Story Genre": "story_genre_1",
            "Channel Language": "channel_language",
            "Program Name": "pgm_name",
            "NST-Program": "NST-Program",
            "NST-Clip-Start": "NST-Clip-Start",
            "NST-Program End": "NST-Program End",
            "NST-Clip-End": "NST-Clip-End",
            "Duration (HH:MM:SS)": "duration",
            "Duration (seconds)": "duration_seconds",
            "Geography": "geography",
            "Logistics": "logistics",
            "Split": "split",
            "Story Format": "Story_Format",
            "Event/Conclave": "Event",
            "LIVE status": "LiveStatus",
            "Exclusive": "exclusive",
            "Telecast Format": "telecast_format",
            "Personality": "personality",
            "Guest": "guest",
            "Anchor": "anchor",
            "Reporter": "reporter",
            "Cluster": "Cluster"
        }


        df.rename(columns={k: v for k, v in column_mapping.items() if k in df.columns}, inplace=True)
        # df = df.drop(['Sr No', 'NST-Program','NST-Clip-Start','NST-Program End','NST-Clip-End'], axis=1)

        df = df.drop(['NST-Program','NST-Clip-Start','NST-Program End','NST-Clip-End','exclusive','channel_language','week_day'], axis=1)

        df['title']=np.nan
        df['story_genre_2']=np.nan
        df['grap_type']=np.nan
        df['assist_used']=np.nan
        df['BARC_CHN_ID']=np.nan
        new_order = [
            'channel', 'BARC_CHNCD','BARC_CHN_ID', 'Story', 'Sub_Story', 'story_genre_1',
            'story_genre_2', 'pgm_name', 'Pgm_Start_Time', 'Pgm_End_Time', 'clip_start_time',
            'clip_end_time', 'pgm_date', 'week', 'geography', 'title', 'grap_type', 'duration',
            'duration_seconds', 'personality', 'guest', 'anchor', 'reporter', 'logistics',
            'telecast_format', 'assist_used', 'split', 'Story_Format'
        ]
        all_files = []

        for x in df['pgm_date'].unique():
            new_df = df[df['pgm_date'] == x]
            file_name = f'output_{x.replace("/", "_")}.csv'
            new_df.to_csv(file_name, index=False)
            all_files.append(file_name)

        zip_file = create_zip('converted_files.zip', all_files)
        
        # Cleanup: delete the individual CSV files after zipping
        for file in all_files:
            os.remove(file)
        
        # Ensure the zip file exists before attempting to send it
        if os.path.exists(zip_file):
            return send_file(zip_file, as_attachment=True, download_name='converted_files.zip', mimetype='application/zip')
        else:
            return "Error: Zip file was not created", 500


    return render_template('excel_to_csv.html')


@val_route.route('/standardize/', methods=['GET', 'POST'])
def standardize():
    if request.method == "POST":
        file = request.files['file']
        df = pd.read_excel(file, sheet_name="expressreport")
        device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

        model = SentenceTransformer('all-MiniLM-L6-v2').to(device)

        sentences = df['Story'].tolist()

        sentence_embeddings = model.encode(sentences, convert_to_tensor=True, device=device)

        sentence_embeddings = torch.nn.functional.normalize(sentence_embeddings, p=2, dim=1)

        similarity_matrix = torch.mm(sentence_embeddings, sentence_embeddings.T)
        similarity_matrix_np = similarity_matrix.cpu().numpy()
        threshold_low = 0.75
        threshold_high = 1

        mask = (similarity_matrix_np > threshold_low) & (similarity_matrix_np < threshold_high)

        new_list = [[] for _ in range(len(sentences))]
        for i in range(len(sentences)):
            matches = np.where(mask[i, i+1:])[0] + (i + 1)
            new_list[i].extend(matches)

        unique_standardization_set = set()
        for x in new_list:
            stand_set = frozenset(sentences[y] for y in x)
            if len(stand_set) > 1: 
                unique_standardization_set.add(stand_set)

        standardization_list = [list(f_set) for f_set in unique_standardization_set]

        # Convert the standardization list into a DataFrame with cell-by-cell placement
        max_len = max([len(x) for x in standardization_list], default=0)
        df_output = pd.DataFrame([x + [''] * (max_len - len(x)) for x in standardization_list])

        # Save the DataFrame to an Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df_output.to_excel(writer, index=False, header=False)
            writer.close()  # Finalize the writer

        output.seek(0)

        # Send the Excel file as a downloadable response
        return send_file(output, download_name="standardized_output.xlsx", as_attachment=True)

    return render_template('33_validations.html')



@val_route.route('/drift', methods=['GET', 'POST'])
def drift_validations():
    if request.method == 'POST':
        file = request.files['file']
        df = pd.read_excel(file, sheet_name="expressreport")

        df = df.sort_values(by=['Channel Name', 'Program Date', 'Clip Start Time'], ascending=[True, True, True]).reset_index(drop=True)

        if len(df.columns) != 34:
            return "Keep the columns default format (34 columns), Remove extra columns", 400

        # Check if there are any blanks in "Clip Start Time"
        if df["Clip Start Time"].isna().any() or (df["Clip Start Time"] == "").any():
            return "Blank exists in Clip Start Time", 400

        # Check if there are any blanks in "Clip End Time"
        if df["Clip End Time"].isna().any() or (df["Clip End Time"] == "").any():
            return "Blank exists in Clip End Time", 400
        
        # Check if any column contains float values or incorrect format
        if df["Clip Start Time"].apply(
                lambda x: not isinstance(x, str) and not isinstance(x, pd.Timestamp)).any() or \
                df["Clip End Time"].apply(
                    lambda x: not isinstance(x, str) and not isinstance(x, pd.Timestamp)).any():
            return "Clip Start Time or Clip End Time contains invalid format", 400

        df["Clip Start Time"] = pd.to_datetime(df["Clip Start Time"], errors='coerce')
        df["Clip End Time"] = pd.to_datetime(df["Clip End Time"], errors='coerce')

        # Replace Clip Start Time if hour is 00 or 04
        df.loc[df["Clip End Time"].dt.hour.isin([0, 4]), "Clip End Time"] = pd.Timestamp("23:59:59")

        for x in range(len(df) - 1):
            # Drift
            # time_diff = (df.iloc[x + 1]["Clip Start Time"] - df.iloc[x]["Clip End Time"]).total_seconds()
            # if time_diff <= 2.0 and 0 < time_diff:
            #     df.at[x + 1, "Clip Start Time"] = df.at[x, "Clip End Time"]
            # Overlap
            time_diff1 = (df.iloc[x]["Clip End Time"] - df.iloc[x + 1]["Clip Start Time"]).total_seconds()
            if time_diff1 <= 3.0 and 0 < time_diff1:
                df.at[x + 1, "Clip Start Time"] = df.at[x, "Clip End Time"]

        lst = []
        lst1 = []
        for x in range(len(df)):
            print(df.iloc[x]["Clip End Time"], df.iloc[x]["Clip Start Time"])
            lst.append(df.iloc[x]["Clip End Time"] - df.iloc[x]["Clip Start Time"])
            lst1.append(int((df.iloc[x]["Clip End Time"] - df.iloc[x]["Clip Start Time"]).total_seconds()))

        df['new_duration_hours'] = lst
        df['new_duration_seconds'] = lst1
        df['new_time'] = df['new_duration_hours'].dt.total_seconds().apply(
            lambda x: f'{int(x // 3600):02}:{int((x % 3600) // 60):02}:{int(x % 60):02}')
        df['Duration (HH:MM:SS)'] = df['new_time']
        df['Duration (seconds)'] = df['new_duration_seconds']
        df["Clip Start Time"] = df["Clip Start Time"].dt.strftime('%H:%M:%S')
        df["Clip End Time"] = df["Clip End Time"].dt.strftime('%H:%M:%S')
        df['Duration (HH:MM:SS)'] = df['Duration (HH:MM:SS)'].apply(
        lambda x: str(x).split()[2] if len(str(x).split()) > 1 else str(x)
    )
        def string_to_timedelta(time_str):
            h, m, s = map(int, time_str.split(':'))
            return timedelta(hours=h, minutes=m, seconds=s)
        df["Duration (HH:MM:SS)"] = df["Duration (HH:MM:SS)"].apply(string_to_timedelta)
        df.drop(columns=['new_duration_hours', 'new_duration_seconds', 'new_time'], inplace=True)
        # drift correction end
        df["Duration (HH:MM:SS)"] = df["Duration (HH:MM:SS)"].astype(str).str.split("\t").str[-1].str.strip()
        df["Duration (HH:MM:SS)"] = pd.to_timedelta(df["Duration (HH:MM:SS)"])

        # Save the DataFrame to an Excel file with a header
        wb = load_workbook('files\\header2.xlsx')
        sheet_name = 'expressreport'
        make_file(wb, 'expressreport', df, False)
        output = BytesIO()
        # time_style = NamedStyle(name="time_style", number_format="hh:mm:ss")
        # ws = wb.active
        # for row in ws.iter_rows(min_row=2, min_col=21, max_col=21):  # Assuming Duration is column 20
        #     for cell in row:
        #         cell.style = time_style
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='expressreport.xlsx', as_attachment=True)

    return render_template('33_validations.html')


@val_route.route('/overlap', methods=['GET', 'POST'])
def verify_overlap():
    if request.method == 'POST':
        file = request.files['file']
        df = pd.read_excel(file, sheet_name="expressreport")

        df = df.sort_values(by=['Channel Name', 'Program Date', 'Clip Start Time'],
                            ascending=[True, True, True]).reset_index(drop=True)

        if len(df.columns) != 34:
            return "Keep the columns default format (34 columns), Remove extra columns", 400

        # Check if there are any blanks in "Clip Start Time"
        if df["Clip Start Time"].isna().any() or (df["Clip Start Time"] == "").any():
            return "Blank exists in Clip Start Time", 400

        # Check if there are any blanks in "Clip End Time"
        if df["Clip End Time"].isna().any() or (df["Clip End Time"] == "").any():
            return "Blank exists in Clip End Time", 400

        # Check if any column contains float values or incorrect format
        if df["Clip Start Time"].apply(
                lambda x: not isinstance(x, str) and not isinstance(x, pd.Timestamp)).any() or \
                df["Clip End Time"].apply(
                    lambda x: not isinstance(x, str) and not isinstance(x, pd.Timestamp)).any():
            return "Clip Start Time or Clip End Time contains invalid format", 400

        df["Clip Start Time"] = pd.to_datetime(df["Clip Start Time"], errors='coerce')
        df["Clip End Time"] = pd.to_datetime(df["Clip End Time"], errors='coerce')

        # Replace Clip Start Time if hour is 00 or 04
        df.loc[df["Clip End Time"].dt.hour.isin([0, 4]), "Clip End Time"] = pd.Timestamp("23:59:59")

        issues = []

        # check overlap
        for x in range(len(df) - 1):
            # Drift
            # time_diff = (df.iloc[x + 1]["Clip Start Time"] - df.iloc[x]["Clip End Time"]).total_seconds()
            # if time_diff <= 2.0 and 0 < time_diff:
            #     df.at[x + 1, "Clip Start Time"] = df.at[x, "Clip End Time"]
            # Overlap
            time_diff1 = (df.iloc[x]["Clip End Time"] - df.iloc[x + 1]["Clip Start Time"]).total_seconds()
            if time_diff1 <= 3.0 and 0 < time_diff1:
                # df.at[x + 1, "Clip Start Time"] = df.at[x, "Clip End Time"]
                #return "overlap found", 400
                issues.append("overlap found")
                break

        # check duplicate
        duplicates = df.duplicated(subset=['Channel Code', 'Program Date', 'Clip Start Time']).sum()
        if duplicates > 0:
            issues.append('duplicate found')


        # check format
        import re

        # Define regex pattern for HH:MM:SS
        pattern = re.compile(r'^\d{2}:\d{2}:\d{2}$')

        # Check if each value is a string and matches the pattern
        matches = df['Clip Start Time'].map(lambda x: isinstance(x, str) and bool(pattern.match(x))).sum()
        if matches > 0:
            issues.append(f"Clip Start Time Format : {matches}")

        matches = df['Clip End Time'].map(lambda x: isinstance(x, str) and bool(pattern.match(x))).sum()
        if matches > 0:
            issues.append(f"Clip Start Time Format : {matches}")

        if len(issues) > 0:
            return f'{" ,".join(issues)}', 400
        else:
            return "FILE IS GOOD TO GO !!!"

        lst = []
        lst1 = []
        for x in range(len(df)):
            print(df.iloc[x]["Clip End Time"], df.iloc[x]["Clip Start Time"])
            lst.append(df.iloc[x]["Clip End Time"] - df.iloc[x]["Clip Start Time"])
            lst1.append(int((df.iloc[x]["Clip End Time"] - df.iloc[x]["Clip Start Time"]).total_seconds()))

        df['new_duration_hours'] = lst
        df['new_duration_seconds'] = lst1
        df['new_time'] = df['new_duration_hours'].dt.total_seconds().apply(
            lambda x: f'{int(x // 3600):02}:{int((x % 3600) // 60):02}:{int(x % 60):02}')
        df['Duration (HH:MM:SS)'] = df['new_time']
        df['Duration (seconds)'] = df['new_duration_seconds']
        df["Clip Start Time"] = df["Clip Start Time"].dt.strftime('%H:%M:%S')
        df["Clip End Time"] = df["Clip End Time"].dt.strftime('%H:%M:%S')
        df['Duration (HH:MM:SS)'] = df['Duration (HH:MM:SS)'].apply(
            lambda x: str(x).split()[2] if len(str(x).split()) > 1 else str(x)
        )

        def string_to_timedelta(time_str):
            h, m, s = map(int, time_str.split(':'))
            return timedelta(hours=h, minutes=m, seconds=s)

        df["Duration (HH:MM:SS)"] = df["Duration (HH:MM:SS)"].apply(string_to_timedelta)
        df.drop(columns=['new_duration_hours', 'new_duration_seconds', 'new_time'], inplace=True)
        # drift correction end
        df["Duration (HH:MM:SS)"] = df["Duration (HH:MM:SS)"].astype(str).str.split("\t").str[-1].str.strip()
        df["Duration (HH:MM:SS)"] = pd.to_timedelta(df["Duration (HH:MM:SS)"])

        # Save the DataFrame to an Excel file with a header
        wb = load_workbook('files\\header2.xlsx')
        sheet_name = 'expressreport'
        make_file(wb, 'expressreport', df, False)
        output = BytesIO()
        # time_style = NamedStyle(name="time_style", number_format="hh:mm:ss")
        # ws = wb.active
        # for row in ws.iter_rows(min_row=2, min_col=21, max_col=21):  # Assuming Duration is column 20
        #     for cell in row:
        #         cell.style = time_style
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='expressreport.xlsx', as_attachment=True)

    return render_template('33_validations.html')