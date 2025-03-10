from flask import Flask,  Blueprint, request, render_template, send_file
import pandas as pd
import os
import xlrd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO

val_route = Blueprint('val_route', __name__)

@val_route.route('/validations', methods=['GET', 'POST'])
def validations():
    if request.method == 'POST':
        file = request.files['file']
        df = pd.read_excel(file, sheet_name="expressreport")
        df.loc[((df['Story']=='ANCHOR INTRODUCTION') & (df['Telecast Format'].str.lower()!='shows')),'Telecast Format'] = 'Shows'
        GA1 = df[(df['Story Format']=='Debate or Discussion')& (df['Guest'].isna() | df['Anchor'].isna())]
        PA = df[(df['Story Format']=='Interview')& (df['Personality'].isna() | df['Anchor'].isna())]
        A = df[(df['Telecast Format']=='Shows') & (df['Anchor'].isna())]
        GA2 = df[(df['Story Format']=='Telephonic Conversation')& (df['Guest'].isna() | df['Anchor'].isna())]
        GA = pd.concat([GA1,GA2])
        grouped = df.groupby('Story')['Story Genre'].agg(list)
        duplicates= grouped[grouped.apply(lambda x: len(set(x))>1)]
        group = df.groupby('Story')['Story Genre'].agg(set)
        dupli= group[group.apply(lambda x: len(set(x))>1)]
        new_df = pd.DataFrame(dupli).reset_index()
        lst=[]
        count_dict={}

        for i,x in enumerate(duplicates):
            for y in set(x):
                lst.append(f"{y} : {x.count(y)}")
                count_dict[y] = x.count(y)
            new_df.loc[i,'Story Genre Count'] = ' '.join(lst)
            new_df.loc[i, 'Story Genre']= max(count_dict, key=count_dict.get)
            lst.clear()
            count_dict.clear()

        new_df.drop('Story Genre',axis=1,  inplace=True)
        geography = df.groupby('Sub-Story')['Geography'].agg(list)
        geog_duplicates= geography[geography.apply(lambda x: len(set(x))>1)]
        geog = df.groupby('Sub-Story')['Geography'].agg(set)
        geog_dup= geog[geog.apply(lambda x: len(set(x))>1)]
        geog_df = pd.DataFrame(geog_dup).reset_index()
        lst2=[]
        for i,x in enumerate(geog_duplicates):

            for y in set(x):

                # print(y , x.count(y))
                lst2.append(f"{y} : {x.count(y)}")
            geog_df.loc[i,'Geography Count'] = ' '.join(lst2)
            lst2.clear()
        geog_df.drop('Geography',axis=1,  inplace=True)
        # excel_writer = pd.ExcelWriter('validations.xlsx', engine='openpyxl')
        # new_df.to_excel(excel_writer, sheet_name='expressreport', index=False)
        # new_df.to_excel(excel_writer, sheet_name='Multiple Story Genre', index=False)
        # geog_df.to_excel(excel_writer, sheet_name='Multiple Geography', index=False)
        # GA.to_excel(excel_writer, sheet_name='Guest Anchor Nulls', index=False)
        # A.to_excel(excel_writer, sheet_name='Anchor Nulls', index=False)
        # PA.to_excel(excel_writer, sheet_name='Anchor Personality Nulls', index=False)
        # excel_writer.close()
        wb = load_workbook('files\\header.xlsx')
        sheet_name = 'expressreport' 
        make_file(wb, 'expressreport', df, False)
        make_file(wb, 'Multiple Story Genre', new_df, True)
        make_file(wb, 'Multiple Geography', geog_df, True)
        make_file(wb, 'Guest Anchor Nulls', GA, True)
        make_file(wb, 'Anchor Nulls', A, True)
        make_file(wb, 'Anchor Personality Nulls', PA, True)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output,download_name='expressreport.xlsx', as_attachment=True)
    return render_template('33_validations.html')


def make_file(wb, sheet_name, df, is_header):    
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        next_row = ws.max_row + 1
    else:
        ws = wb[sheet_name]
        next_row = ws.max_row + 1
    # Append the new data to the selected sheet
    if is_header:
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=is_header), 0):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=next_row+r_idx-1, column=c_idx, value=value)
    else:
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=is_header), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=next_row+r_idx-1, column=c_idx, value=value)