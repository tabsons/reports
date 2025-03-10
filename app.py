import requests
import pandas as pd
import numpy as np
from difflib import SequenceMatcher
from flask import Flask, render_template, request, send_file, redirect, session
from datetime import datetime, time as t, timedelta
import re
import pytz
import ren
import time as time
st = time.time()

from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

from routes.celebs import celebs_route
from routes.xen import xen_route
from routes.story import story_routes
from routes.eq import eq_route
from routes.validators import val_route





app = Flask(__name__)
app.register_blueprint(celebs_route)
app.register_blueprint(xen_route)
app.register_blueprint(story_routes)
app.register_blueprint(eq_route)
app.register_blueprint(val_route)

app.secret_key = '10.18.0.26'

# Dictionary to store the valid usernames and passwords
valid_users = {
    'admin': 'password',
    'Chirag': 'Chirag@123',
    'user2': 'password2',
}


@app.route('/home')
def homepage():
    return render_template('login.html')

@app.route('/testing')
def test():
    return render_template('test.html')


@app.route('/')
def index():
    # Check if the user is logged in
    if 'username' in session:
        return render_template('home.html')
    else:
        return redirect('/login')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # Process the login form data
        username = request.form.get('username')
        password = request.form.get('password')

        # Perform authentication logic here
        # Check if the username and password are valid
        if username in valid_users and valid_users[username] == password:
            # Authentication successful, store the username in session
            session['username'] = username
            return redirect('/')
        else:
            # Authentication failed
            return render_template('login.html', error='Invalid credentials')

    # If the request method is GET, render the login form
    return render_template('login.html')

@app.route('/logout')
def logout():
    # Clear the session and redirect to the login page
    session.clear()
    return redirect('/login')



@app.route('/download/<path:filename>', methods=['GET'])
def download(filename):
    return send_file(filename, as_attachment=True)

@app.route('/shakaplayer', methods=['GET'])
def shaka_player():
    return render_template('shaka_player.html')



@app.route('/newUrl', methods=['GET'])
def newUrl():
    return render_template('33_validations.html')


# Route to render the upload page
@app.route('/pivot')
def upload_file():
    return render_template('index_process.html')


# Route to process the uploaded file
@app.route('/process33', methods=['POST'])
def process_file():
    if 'file' not in request.files:
        return "No file uploaded", 400

    file = request.files['file']
    if not file:
        return "File not selected", 400

    # Load the Excel file into a DataFrame
    df_pandas = pd.read_excel(file)
    if df_pandas.shape[1] < 33:
        return "Please, Correct the file. Less than 33 fields.", 400
    elif df_pandas.shape[1] > 35:
        return "Please, Correct the file. More than 35 fields.", 400
    else:
        pass
    if (df_pandas.notna().sum() > 1).any():
        pass
    else:
        return "Please, Correct the file. No fields.", 400
    if 'Sub-Story Genre' in df_pandas.columns:
        df = df_pandas.drop(columns=['Sub-Story Genre'])
    else:
        df = df_pandas
    # Check the column type first
    # Convert values of type `datetime.time` to strings, then apply `to_timedelta`
    df['Duration (HH:MM:SS)'] = df['Duration (HH:MM:SS)'].apply(
        lambda x: x.strftime('%H:%M:%S') if isinstance(x, t) else x
    )
    df['Duration (HH:MM:SS)'] = pd.to_timedelta(df['Duration (HH:MM:SS)'], errors='coerce')

    df['Duration (HH:MM:SS)'] = df['Duration (HH:MM:SS)'].apply(
        lambda x: str(x).split()[2] if len(str(x).split()) > 1 else str(x)
    )

    # Add 'Status' column
    df['Status'] = df.apply(
        lambda row: (
            'Both Story and Substory empty' if pd.isna(row['Story']) and pd.isna(row['Sub-Story']) else
            'Story empty' if pd.isna(row['Story']) else
            'Substory empty' if pd.isna(row['Sub-Story']) else
            ''
        ),
        axis=1
    )

    # Convert durations to seconds
    def convert_to_seconds(duration):
        h, m, s = map(int, duration.split(':'))
        return h * 3600 + m * 60 + s

    df['Duration (Seconds)'] = df['Duration (HH:MM:SS)'].apply(convert_to_seconds)

    # Create a pivot table
    pivot_df = pd.pivot_table(df, values='Duration (Seconds)',
                              index='Channel Name', columns='Program Date',
                              aggfunc='sum', fill_value=0)

    # Convert seconds back to HH:MM:SS
    def seconds_to_hms(seconds):
        hours = seconds // 3600
        minutes = (seconds % 3600) // 60
        seconds = seconds % 60
        return f'{hours:02}:{minutes:02}:{seconds:02}'

    # pivot_df['Duration (HH:MM:SS)'] = pivot_df['Duration (Seconds)'].apply(seconds_to_hms)
    # pivot_df = pivot_df.drop(columns=['Duration (Seconds)'])
    pivot_df = pivot_df.applymap(seconds_to_hms)
    df['key'] = df['Program Date'].astype(str) + '_' + df['Channel Code'].astype(str) + '_' + df[
        'Clip Start Time'].astype(str)
    df['key'] = df['key'].where(df['key'].duplicated(keep=False), 0)
    # Generate week number and filename
    first_entry = df['Program Date'].iloc[0]
    date_object = datetime.strptime(first_entry, "%d/%m/%Y")
    # week_number = (date_object - timedelta(days=date_object.weekday() + 2)).strftime('%U')
    # week_number = int(week_number) + 2
    week = df["Week"][0]
    file_name = f"{week} {date_object.year}{date_object.month:02}{date_object.day:02} 33.xlsx"
    df['Story Genre'] = df['Story Genre'].replace({
        'CAREER & EDUCATION': 'CAREER/EDUCATION',
        'SPIRITUAL & RELIGION': 'SPIRITUAL/RELIGION',
        'TECHNOLOGY & GADGET': 'TECHNOLOGY/GADGET'
    })

    def string_to_timedelta(time_str):
        h, m, s = map(int, time_str.split(':'))
        return timedelta(hours=h, minutes=m, seconds=s)

    df["Duration (HH:MM:SS)"] = df["Duration (HH:MM:SS)"].apply(string_to_timedelta)
    # pivot_df['Duration (HH:MM:SS)'] = pivot_df['Duration (HH:MM:SS)'].apply(string_to_timedelta)
    pivot_df = pivot_df.applymap(string_to_timedelta)
    df = df.drop(columns=['Duration (Seconds)'])
    df = df[df['Duration (seconds)'] != 0]
    df = df.sort_values(by=['Channel Name', 'Program Date', 'Clip Start Time'])
    # Save to Excel
    output_path = f'uploads/{file_name}'
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='expressreport')
        pivot_df.to_excel(writer, sheet_name='pivot')

        workbook = writer.book
        worksheet = writer.sheets['expressreport']
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        # Apply yellow fill based on 'Status'
        for row_idx, row in enumerate(df.itertuples(), start=2):
            if row.Status:
                for col_idx in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = yellow_fill
            elif row.key != 0:
                for col_idx in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = red_fill
        ##
        grey_fill = PatternFill(start_color="6f6f6f", end_color="6f6f6f", fill_type="solid")
        # Define a white font style
        white_font = Font(color="FFFFFF", bold=True)  # White text with bold
        # Assuming `worksheet` is already defined, and `df` is the DataFrame

        # Apply grey fill to the first row (header)
        for col_idx in range(1, len(df.columns) + 1):  # Iterate over each column
            cell = worksheet.cell(row=1, column=col_idx)  # Row 1 is the header
            cell.fill = grey_fill  # Apply grey fill to the header
            cell.font = white_font

    # Step 4: Apply Time Formatting with OpenPyXL
    # Load the Excel file
    wb = load_workbook(output_path)

    # Define the time style for Excel (hh:mm:ss) once
    time_style = NamedStyle(name="time_style", number_format="hh:mm:ss")

    # Apply the time style to the first sheet (active sheet)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, min_col=20, max_col=20):  # Assuming Duration is column 20
        for cell in row:
            cell.style = time_style
    ws.delete_cols(35)
    ws.delete_cols(35)
    # Apply the time style to the second sheet (Sheet2)
    ws = wb["pivot"]  # Replace "Sheet2" with the actual name if needed
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=ws.max_column):  # Assuming Duration is column 3 (C)
        for cell in row:
            cell.style = time_style
    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1,
                            max_col=ws.max_column):  # Assuming Duration is column 3 (C)
        for cell in row:
            cell.fill = grey_fill  # Apply grey fill to the header
            cell.font = white_font
    # Save the changes
    wb.save(output_path)

    return send_file(output_path, as_attachment=True)




if __name__ == '__main__':
    app.run(debug=True,host='0.0.0.0')

